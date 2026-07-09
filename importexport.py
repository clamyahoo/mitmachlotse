"""
Import- und Exportfunktionen für die Projekttage-App.
"""

import csv
import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from odf.opendocument import load as odf_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False
import database as db


TEILNEHMER_FELDER_BASIS = [
    ("nachname",  "Nachname"),
    ("vorname",   "Vorname"),
    ("ganzer_name_kombiniert", "Ganzer Name in einer Spalte (Spaltentitel \"Name\")"),
    ("stufe",      "Gruppenbereich"),
    ("stufenzusatz", "Gruppenzusatz (Kürzel)"),
    ("klasse_kombiniert", "Gruppenbereich + Gruppenzusatz kombiniert"),
    # "geschlecht" wurde entfernt – optional über Zusatzfelder konfigurierbar
    ("wunsch_1",  "Wunsch 1"),
    ("wunsch_2",  "Wunsch 2"),
    ("wunsch_3",  "Wunsch 3"),
    ("wunsch_4",  "Wunsch 4"),
    ("wunsch_5",  "Wunsch 5"),
    ("projekt",   "Projekt"),
]


def get_schueler_felder(max_wuensche: int = None) -> list:
    """
    Gibt die vollständige Feldliste für den Import zurück.
    Wunschfelder werden auf max_wuensche begrenzt (aus Konfig wenn None).
    """
    konfig = db.get_feldkonfig()
    if max_wuensche is None:
        max_wuensche = konfig.get("max_wuensche", 5)
    felder = []
    for key, label in TEILNEHMER_FELDER_BASIS:
        if key == "stufe":
            felder.append((key,
                f"{konfig.get('stufe_label', 'Gruppenbereich')} "
                f"(nur ganze Zahlen, z.\u202fB. 5, 6, 10)"))
        elif key == "stufenzusatz":
            felder.append((key,
                f"{konfig.get('stufenzusatz_label', 'Gruppenzusatz')} (Kürzel)"))
        elif key == "klasse_kombiniert":
            sl = konfig.get('stufe_label', 'Gruppenbereich')
            zl = konfig.get('stufenzusatz_label', 'Gruppenzusatz')
            felder.append((key,
                f"{sl} + {zl} kombiniert "
                f"(wird beim Import automatisch in 2 Felder aufgetrennt; "
                f"im {sl}-Anteil sind nur ganze Zahlen zul\u00e4ssig)"))
        elif key == "ganzer_name_kombiniert":
            felder.append((key, "Ganzer Name in einer Spalte (Spaltentitel \"Name\")"))
        elif key == "projekt":
            felder.append((key, konfig.get("projekt_label", label)))
        elif key.startswith("wunsch_"):
            nr = int(key.split("_")[1])
            if nr <= max_wuensche:
                felder.append((key, label))
            # Wünsche über max_wuensche werden nicht angeboten
        else:
            felder.append((key, label))
    for i in range(1, 4):
        lbl = konfig.get(f"extra_{i}_label", "")
        if lbl:
            felder.append((f"extra_{i}", lbl))
    return felder


# Abwärtskompatibilität: SCHUELER_FELDER als dynamische Property
def _get_schueler_felder_compat():
    return get_schueler_felder()
SCHUELER_FELDER = property(_get_schueler_felder_compat)

# Felder, die beim Import aufgesplittet werden, statt direkt in die
# Datenbank zu gehen
TEILNEHMER_VIRTUELLES_NAMENSFELD = "ganzer_name_kombiniert"
TEILNEHMER_VIRTUELLES_KLASSENFELD = "klasse_kombiniert"

def detect_wunsch_anzahl(headers: list) -> int:
    """
    Zählt wie viele Wunschspalten (W1, Wunsch 1, wunsch1 …) in einem
    Header-Set erkannt werden. Gibt 0 zurück wenn keine gefunden.
    """
    patterns = [
        ["wunsch 1", "wunsch1", "w1"],
        ["wunsch 2", "wunsch2", "w2"],
        ["wunsch 3", "wunsch3", "w3"],
        ["wunsch 4", "wunsch4", "w4"],
        ["wunsch 5", "wunsch5", "w5"],
    ]
    headers_lower = [h.lower().strip() for h in headers]
    count = 0
    for i, aliases in enumerate(patterns):
        if any(a in headers_lower for a in aliases):
            count = i + 1  # fortlaufend – letzter Fund = Anzahl
    return count
    """Gibt Projektfelder mit konfigurierten Labels zurück."""
    k = db.get_feldkonfig()
    pl = k.get("projekt_label", "Projekt")
    sl = k.get("stufe_label",   "Gruppenbereich")
    return [
        ("nummer",      "Nummer"),
        ("projektname", db.get_label_formen(pl)["name"]),
        ("stufenmin",   f"{sl} min"),
        ("stufenmax",   f"{sl} max"),
        ("tnmin",       "Plätze min"),
        ("tnmax",       "Plätze max"),
    ]


# Abwärtskompatibilität
PROJEKT_FELDER = [
    ("nummer",      "Nummer"),
    ("projektname", "Projektname"),
    ("stufenmin",   "Gruppenbereich min"),
    ("stufenmax",   "Gruppenbereich max"),
    ("tnmin",       "Plätze min"),
    ("tnmax",       "Plätze max"),
]


def get_projekt_felder() -> list:
    """Gibt Projektfelder mit konfigurierten Labels zurück."""
    k  = db.get_feldkonfig()
    pl = k.get("projekt_label", "Option")
    sl = k.get("stufe_label",   "Gruppenbereich")
    ll = k.get("leitung_label", "").strip()
    # Leitung wird beim Import IMMER als App-Feld angeboten (zwischen
    # Nummer und Optionsname), auch wenn die Spalte über "Spaltenbezeichnungen
    # anpassen" noch nicht aktiviert wurde. Wird ihr keine Quelle zugeordnet,
    # bleibt sie einfach weg; wird ihr eine Quelle zugeordnet, fragt der
    # Import-Dialog nach, ob die Spalte jetzt aktiviert werden soll
    # (siehe ImportDialog._pruefe_leitungsspalte).
    felder = [
        ("nummer",  "Nummer"),
        ("leitung", ll if ll else "Leitung"),
        ("projektname", db.get_label_formen(pl)["name"]),
    ]
    felder += [
        ("stufenmin",   f"{sl} min"),
        ("stufenmax",   f"{sl} max"),
        ("tnmin",       "Plätze min"),
        ("tnmax",       "Plätze max"),
    ]
    for i in range(1, 4):
        lbl = k.get(f"projekt_extra_{i}_label", "")
        if lbl:
            felder.append((f"extra_{i}", lbl))
    return felder


def detect_csv_separator(filepath: str) -> str:
    """
    Versucht das Trennzeichen einer CSV-Datei automatisch zu erkennen.
    Gibt ';', ',' oder '\t' zurück.
    """
    try:
        with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
            sample = f.read(4096)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        return dialect.delimiter
    except Exception:
        # Fallback: Semikolon zählen vs. Komma zählen
        try:
            with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
                first = f.readline()
            if first.count(";") >= first.count(","):
                return ";"
            return ","
        except Exception:
            return ";"


def _default_teilnehmer():
    return {
        "nachname": "-", "vorname": "-", "stufe": "-",
        "stufenzusatz": "-", "geschlecht": "-",
        "wunsch_1": 0, "wunsch_2": 0, "wunsch_3": 0,
        "wunsch_4": 0, "wunsch_5": 0, "projekt": 0,
    }


def _default_projekt():
    """
    Startwerte für den Import. Wird ein Feld nicht zugeordnet
    ("nicht importieren"), bleibt dieser Wert stehen. Bei den
    Pflichtfeldern Gruppenbereich/Plätze min+max wird bewusst
    durchgängig 0 verwendet (statt z. B. asymmetrisch min=0/max=30) —
    ein nicht importierter Bereich soll nicht wie ein plausibler,
    absichtlich gesetzter Bereich aussehen, sondern klar erkennbar auf
    Nachbearbeitung hinweisen.
    """
    return {
        "nummer": 0, "projektname": "-", "leitung": "",
        "stufenmin": 0, "stufenmax": 0, "tnmin": 0, "tnmax": 0,
        "extra_1": "", "extra_2": "", "extra_3": "",
    }


def _coerce_int(val, default=0):
    """
    Wandelt val in int um. Excel/LibreOffice speichern reine
    Zahlenspalten nach dem Bearbeiten oft als Kommazahl ("1.0" statt "1")
    — ohne Sonderbehandlung würde int("1.0") eine ValueError auslösen und
    lautlos auf `default` zurückfallen (z. B. Nummer 0 → Zeile wird beim
    Import stillschweigend verworfen). Daher zusätzlich ein Fallback über
    float().
    """
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return int(float(str(val).replace(",", ".")))
        except (ValueError, TypeError):
            return default


def _normalisiere_jgst(val) -> str:
    """
    Normalisiert eine Jahrgangsstufen-Angabe zu einem sauberen String.

    Excel/LibreOffice liefern Zahlenfelder beim Import oft als Float
    ("5.0" statt "5"), was sonst zu Fehlern bei der Jgst-Auswertung führen
    kann (z. B. beim Algorithmus-Abgleich gegen stufenmin/stufenmax von
    Projekten). Beispiele:
        "5.0"  -> "5"
        "5,0"  -> "5"
        "5"    -> "5"
        "5a"   -> "5a"   (Klassenbezeichnung bleibt unangetastet)
        "-"    -> "-"
        ""     -> "-"
    """
    s = str(val).strip()
    if not s or s == "-":
        return "-"
    # Nur bereinigen, wenn es sich um eine reine (Dezimal-)Zahl handelt,
    # z. B. "5.0" oder "5,0" -- gemischte Werte wie "5a" unverändert lassen
    bereinigt = s.replace(",", ".")
    try:
        zahl = float(bereinigt)
        if zahl == int(zahl):
            return str(int(zahl))
        return s  # z. B. "5.5" ergibt keinen Sinn als Jgst -> unverändert
    except ValueError:
        return s


# ── CSV lesen ────────────────────────────────────────────────────────────────

def _detect_encoding(filepath: str) -> str:
    """Probiert gängige Encodings durch und gibt das erste funktionierende zurück."""
    candidates = ["utf-8-sig", "utf-8", "windows-1252", "latin-1", "cp1250", "iso-8859-1"]
    with open(filepath, "rb") as f:
        raw = f.read()
    for enc in candidates:
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"  # Fallback: latin-1 akzeptiert alle Bytes


def _generiere_spaltennamen(anzahl: int) -> list:
    """Erzeugt generische Spaltennamen 'Spalte 1', 'Spalte 2', ..."""
    return [f"Spalte {i+1}" for i in range(anzahl)]


def read_csv(filepath: str, delimiter: str = ";", has_header: bool = True) -> tuple[list, list]:
    """Liest CSV und gibt (headers, rows) zurück. Erkennt Encoding automatisch."""
    encoding = _detect_encoding(filepath)
    with open(filepath, newline="", encoding=encoding) as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)
    if not rows:
        return [], []
    if has_header:
        return rows[0], rows[1:]
    else:
        spaltenzahl = max(len(r) for r in rows)
        return _generiere_spaltennamen(spaltenzahl), rows


# ── Excel/Calc lesen ─────────────────────────────────────────────────────────

def read_excel(filepath: str, has_header: bool = True) -> tuple[list, list]:
    """Liest .xlsx/.ods und gibt (headers, rows) zurück."""
    if filepath.lower().endswith(".ods"):
        return read_ods(filepath, has_header)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    data = [[str(cell.value) if cell.value is not None else "" for cell in row]
            for row in ws.iter_rows()]
    if not data:
        return [], []
    if has_header:
        return data[0], data[1:]
    else:
        spaltenzahl = max(len(r) for r in data)
        return _generiere_spaltennamen(spaltenzahl), data


def read_ods(filepath: str, has_header: bool = True) -> tuple[list, list]:
    """Liest .ods (LibreOffice Calc) und gibt (headers, rows) zurück."""
    if not ODF_AVAILABLE:
        raise ImportError(
            "odfpy ist nicht installiert.\n"
            "Bitte ausführen: pip install odfpy"
        )
    doc = odf_load(filepath)
    sheets = doc.spreadsheet.getElementsByType(Table)
    if not sheets:
        return [], []
    ws = sheets[0]  # Erstes Tabellenblatt

    data = []
    for row_el in ws.getElementsByType(TableRow):
        row = []
        for cell_el in row_el.getElementsByType(TableCell):
            # Wiederholungsspalten berücksichtigen
            repeat = int(cell_el.getAttribute("numbercolumnsrepeated") or 1)
            # Zelltext extrahieren
            ps = cell_el.getElementsByType(P)
            val = "".join(str(p) for p in ps) if ps else ""
            row.extend([val] * repeat)
        # Trailing-Leerfelder abschneiden
        while row and row[-1] == "":
            row.pop()
        if row:  # Leerzeilen überspringen
            data.append(row)

    if not data:
        return [], []
    if has_header:
        return data[0], data[1:]
    else:
        spaltenzahl = max(len(r) for r in data)
        return _generiere_spaltennamen(spaltenzahl), data

def _kandidaten_spaltennamen(konfig: dict = None) -> set:
    """
    Menge bekannter (kleingeschriebener) Spaltennamen, anhand derer eine
    plausible Kopfzeile erkannt wird — genutzt, um beim Reimport
    Zier-/Titelzeilen zu überspringen, die der Listenexport oberhalb der
    eigentlichen Tabelle einfügt (Kopfzeile, Gruppen-/Optionsname).
    """
    konfig = konfig if konfig is not None else db.get_feldkonfig()
    sl = konfig.get("stufe_label", "Gruppenbereich").lower()
    zl = konfig.get("stufenzusatz_label", "Gruppenzusatz").lower()
    pl = konfig.get("projekt_label", "Projekt").lower()
    return {
        "name", "nachname", "vorname", "ganzer name", "vollname",
        "klasse", "stufe", "jahrgangsstufe", "jahrgang", "jg", "jgst", "jgst.",
        sl, zl, "klassenzusatz", "stufenzusatz", "zusatz",
        "wunsch 1", "wunsch 2", "wunsch 3", "wunsch 4", "wunsch 5",
        "wunsch1", "wunsch2", "wunsch3", "wunsch4", "wunsch5",
        "w1", "w2", "w3", "w4", "w5",
        "projekt", pl, "projekt-nr.", "option-nr.", f"{pl}-nr.",
        "nummer", "nr", "nr.", "projektname", "titel", f"{pl}name",
        "stufenmin", "stufenmax", "jgst. min", "jgst. max",
        "tnmin", "tnmax", "plätze min", "plätze max",
        "zuteilung", "geschlecht",
        "raumname", "raum", "kapazität", "kapazitaet", "beschreibung",
    }


def _ist_plausible_headerzeile(row: list, kandidaten: set) -> bool:
    """True, wenn mindestens 2 Zellen der Zeile einem bekannten
    App-Spaltennamen entsprechen (Groß-/Kleinschreibung egal)."""
    treffer = 0
    for zelle in row:
        wert = str(zelle).strip().lower().rstrip(".")
        if wert in kandidaten:
            treffer += 1
            if treffer >= 2:
                return True
    return False


def entferne_geisterzeilen(headers: list, rows: list, konfig: dict = None) -> tuple[list, list]:
    """
    Entfernt „Geisterzeilen", die beim Reimport einer selbst exportierten
    Gesamtliste (nach Gruppen ODER nach Optionen) mitten im Datenbereich
    stehen: die je Abschnitt wiederholten Kopfzeilen und die
    Abschnittsüberschriften (Gruppen-/Optionsname). Ohne diese Bereinigung
    würden daraus fehlerhafte Datensätze entstehen (z. B. ein „Teilnehmer"
    namens „1: Holzwerkstatt" oder „Name").

    Greift bewusst nur, wenn die Datei tatsächlich wie ein mehrteiliger
    Selbstexport aussieht (mindestens eine wiederholte Kopfzeile im
    Datenbereich). Für gewöhnliche Importdateien bleibt alles unverändert,
    damit keine legitimen Zeilen verloren gehen.
    """
    if not rows:
        return headers, rows
    kandidaten = _kandidaten_spaltennamen(konfig)
    hat_wiederholten_header = any(
        _ist_plausible_headerzeile(r, kandidaten) for r in rows
    )
    if not hat_wiederholten_header:
        return headers, rows

    bereinigt = []
    for row in rows:
        # (a) wiederholte Kopfzeile (≥2 bekannte Spaltennamen)
        if _ist_plausible_headerzeile(row, kandidaten):
            continue
        # (b) Abschnittsüberschrift (nur erste Spalte gefüllt) oder komplette
        # Leerzeile (Abstandszeile zwischen den Abschnitten). Echte Teilnehmer-/
        # Optionszeilen aus einem Export haben immer noch mindestens
        # Gruppenbereich bzw. weitere Felder gefüllt.
        nicht_leer = [i for i, v in enumerate(row) if str(v).strip() != ""]
        if not nicht_leer or nicht_leer == [0]:
            continue
        bereinigt.append(row)
    return headers, bereinigt


def bereinige_titelzeilen(headers: list, rows: list, konfig: dict = None,
                           max_suchtiefe: int = 15) -> tuple[list, list]:
    """
    Erkennt und entfernt Zier-/Titelzeilen vor der eigentlichen Kopfzeile,
    wie sie z. B. beim eigenen Listenexport entstehen (Kopfzeile +
    Gruppen-/Optionsname stehen dort über der Tabelle). Praktisch für den
    Reimport von Dateien, die zuvor mit "Gesamtliste exportieren" oder
    dem Fenster-Export erzeugt wurden. Zusätzlich werden über
    entferne_geisterzeilen die je Abschnitt wiederholten Kopf- und
    Titelzeilen aus dem Datenbereich entfernt.

    Greift nur, wenn die aktuell erkannten Header NICHT plausibel
    aussehen, aber eine der nächsten Zeilen schon — sonst bleibt alles
    unverändert (defensiv: im Zweifel wird nichts entfernt).
    """
    if not headers:
        return headers, rows
    kandidaten = _kandidaten_spaltennamen(konfig)
    if _ist_plausible_headerzeile(headers, kandidaten):
        return entferne_geisterzeilen(headers, rows, konfig)

    for i, row in enumerate(rows[:max_suchtiefe]):
        if _ist_plausible_headerzeile(row, kandidaten):
            neue_headers = [str(v).strip() for v in row]
            neue_rows = rows[i + 1:]
            return entferne_geisterzeilen(neue_headers, neue_rows, konfig)

    return headers, rows  # keine plausible Kopfzeile gefunden -> unverändert


def merge_import_dateien(filepaths: list) -> tuple[list, list]:
    """
    Liest mehrere Quelldateien (.xlsx, .ods, .csv) ein und führt sie zu
    einem gemeinsamen Header-/Zeilensatz zusammen — z. B. wenn für jede
    Gruppe eine eigene Wunschliste zurückgekommen ist.

    Vor dem Zusammenführen wird pro Datei geprüft, ob ihr die echte
    Kopfzeile z. B. durch eine exportierte Titelzeile vorangestellt ist
    (siehe bereinige_titelzeilen) — so lassen sich auch mit "Gesamtliste
    exportieren" erzeugte Dateien direkt wieder importieren.

    Spalten werden anhand ihres Namens abgeglichen: Die Spaltenreihenfolge
    der ersten Datei ist maßgeblich, zusätzliche Spalten aus späteren
    Dateien werden rechts angehängt. Fehlt eine Spalte in einer Datei,
    wird sie für deren Zeilen mit "" aufgefüllt.

    Gibt (merged_headers, merged_rows) zurück.
    """
    merged_headers: list = []
    merged_rows: list = []
    konfig = db.get_feldkonfig()

    for path in filepaths:
        if path.lower().endswith(".csv"):
            delimiter = detect_csv_separator(path)
            headers, rows = read_csv(path, delimiter, True)
        else:
            headers, rows = read_excel(path, True)  # .xlsx und .ods

        headers, rows = bereinige_titelzeilen(headers, rows, konfig)

        headers = [str(h).strip() for h in headers]
        if not headers:
            continue

        for h in headers:
            if h not in merged_headers:
                merged_headers.append(h)

        spalten_index = [merged_headers.index(h) for h in headers]
        for row in rows:
            neue_zeile = [""] * len(merged_headers)
            for quell_idx, ziel_idx in enumerate(spalten_index):
                if quell_idx < len(row):
                    neue_zeile[ziel_idx] = row[quell_idx]
            merged_rows.append(neue_zeile)

    return merged_headers, merged_rows


def detect_wert_varianten(headers: list, rows: list, spaltenname: str) -> dict:
    """
    Erkennt in einer Spalte Werte mit Klammerzusatz (z. B. "a (MüS)"),
    deren Basisform (Text vor der Klammer) auch als eigenständiger Wert
    in derselben Spalte vorkommt — typischerweise durch einen Zusatz wie
    ein Klassenlehrer-Kürzel, den jemand beim Ausfüllen ergänzt hat.

    Gibt {ursprünglicher_wert: bereinigter_wert} zurück.
    """
    if spaltenname not in headers:
        return {}
    idx = headers.index(spaltenname)

    werte = set()
    for row in rows:
        if idx < len(row):
            wert = str(row[idx]).strip()
            if wert:
                werte.add(wert)

    varianten = {}
    for wert in werte:
        match = re.match(r"^(.*?)\s*\((.*?)\)\s*$", wert)
        if match:
            basis = match.group(1).strip()
            if basis and basis in werte:
                varianten[wert] = basis
    return varianten


def wende_wert_bereinigung_an(headers: list, rows: list,
                               spaltenname: str, ersetzungen: dict) -> None:
    """Wendet {alt: neu}-Ersetzungen in-place auf eine benannte Spalte an."""
    if not ersetzungen or spaltenname not in headers:
        return
    idx = headers.index(spaltenname)
    for row in rows:
        if idx < len(row):
            wert = str(row[idx]).strip()
            if wert in ersetzungen:
                row[idx] = ersetzungen[wert]


def schreibe_merge_temp_xlsx(headers: list, rows: list) -> str:
    """
    Schreibt zusammengeführte Header/Zeilen in eine temporäre .xlsx-Datei
    und gibt deren Pfad zurück. Wird genutzt, um die zusammengeführten
    Daten über den bestehenden Einzeldatei-Importweg (Spaltenzuordnung,
    Wunschanzahl-Erkennung usw.) laufen zu lassen.
    Aufrufer ist für das Löschen der Datei verantwortlich.
    """
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def split_ganzer_name(value: str) -> tuple[str, str]:
    """
    Trennt einen kombinierten Namen in (nachname, vorname) auf.

    Unterstützte Formate:
      "Nachname, Vorname"   -> ("Nachname", "Vorname")
      "Nachname Vorname"    -> ("Nachname", "Vorname")  (letztes Wort = Vorname-Annahme
                                funktioniert NICHT zuverlässig bei Leerzeichen-Format,
                                daher: erstes Wort = Nachname, Rest = Vorname,
                                außer bei Komma-Format)
    Bei "Nachname Vorname" (ohne Komma) wird angenommen:
      erstes Wort = Nachname, restliche Wörter = Vorname.
    Das ist die in deutschen Schulverwaltungsprogrammen gängige Reihenfolge.
    """
    value = value.strip()
    if not value:
        return "-", "-"

    if "," in value:
        teile = value.split(",", 1)
        nachname = teile[0].strip()
        vorname = teile[1].strip() if len(teile) > 1 else "-"
        return (nachname or "-", vorname or "-")

    # Kein Komma: nach Leerzeichen trennen
    teile = value.split(None, 1)
    if len(teile) == 2:
        return (teile[0].strip() or "-", teile[1].strip() or "-")
    elif len(teile) == 1:
        return (teile[0].strip() or "-", "-")
    return "-", "-"


def split_klasse(value: str) -> tuple[str, str]:
    """
    Trennt eine kombinierte Klassenbezeichnung in (jgst, klassenzusatz) auf.

    Beispiele:
      "5a"    -> ("5", "a")
      "10c"   -> ("10", "c")
      "7a1"   -> ("7", "a1")
      "K1"    -> ("-", "K1")   (keine führende Zahl -> alles als Zusatz)
      "5"     -> ("5", "-")    (keine Buchstaben -> kein Zusatz)
    """
    value = value.strip()
    if not value:
        return "-", "-"

    match = re.match(r"^(\d+)\s*(.*)$", value)
    if not match:
        # Keine führende Zahl erkennbar (z. B. Kursbezeichnungen wie "K1")
        return "-", value

    jgst = match.group(1)
    zusatz = match.group(2).strip()
    return (jgst or "-", zusatz or "-")


def import_teilnehmer(headers: list, rows: list, mapping: dict, append: bool = False) -> list:
    """
    mapping = {app_feld: quelltabellen_index, ...}
    append=False löscht vorher alle Schüler.

    Unterstützt sowohl getrennte Felder (nachname, vorname / stufe, stufenzusatz)
    als auch kombinierte Felder (ganzer Name / ganze Klasse), die automatisch
    aufgesplittet werden.

    Gibt die Liste der IDs der neu importierten Teilnehmer/innen zurück.
    """
    if not append:
        db.clear_teilnehmer()

    int_felder = {"wunsch_1", "wunsch_2", "wunsch_3", "wunsch_4", "wunsch_5", "projekt"}
    extra_felder = {"extra_1", "extra_2", "extra_3"}
    virtuelle_felder = {TEILNEHMER_VIRTUELLES_NAMENSFELD, TEILNEHMER_VIRTUELLES_KLASSENFELD}
    namensfeld_idx = mapping.get(TEILNEHMER_VIRTUELLES_NAMENSFELD)
    klassenfeld_idx = mapping.get(TEILNEHMER_VIRTUELLES_KLASSENFELD)

    importierte_ids = []

    for row in rows:
        record = _default_teilnehmer()
        # Extra-Felder mit leerem Standardwert vorbelegen
        for ex in extra_felder:
            record[ex] = ""
        for feld, idx in mapping.items():
            if idx is None or feld in virtuelle_felder:
                continue
            try:
                val = row[idx] if idx < len(row) else ""
            except (IndexError, TypeError):
                val = ""
            if feld in int_felder:
                record[feld] = _coerce_int(val)
            elif feld == "stufe":
                record[feld] = _normalisiere_jgst(val) if val else record[feld]
            else:
                record[feld] = val if val else record[feld]

        # Kombiniertes Namensfeld auswerten (überschreibt nachname/vorname,
        # falls zusätzlich zugeordnet und nicht leer)
        if namensfeld_idx is not None:
            try:
                roh = row[namensfeld_idx] if namensfeld_idx < len(row) else ""
            except (IndexError, TypeError):
                roh = ""
            if roh:
                nachname, vorname = split_ganzer_name(roh)
                record["nachname"] = nachname
                record["vorname"] = vorname

        # Kombiniertes Klassenfeld auswerten (überschreibt jgst/abteilung,
        # falls zusätzlich zugeordnet und nicht leer)
        if klassenfeld_idx is not None:
            try:
                roh = row[klassenfeld_idx] if klassenfeld_idx < len(row) else ""
            except (IndexError, TypeError):
                roh = ""
            if roh:
                jgst, zusatz = split_klasse(roh)
                record["stufe"] = _normalisiere_jgst(jgst)
                record["stufenzusatz"] = zusatz

        # Nur importieren wenn mindestens Nachname vorhanden
        if record["nachname"] and record["nachname"] != "-":
            neue_id = db.insert_teilnehmer(record)
            importierte_ids.append(neue_id)

    return importierte_ids


def get_raum_felder() -> list:
    """App-Felder für den Raumlisten-Import (key, Anzeigelabel)."""
    return [
        ("name",         "Raumname"),
        ("kapazitaet",   "Kapazität"),
        ("beschreibung", "Beschreibung"),
    ]


def _default_raum() -> dict:
    return {"name": "-", "kapazitaet": 0, "beschreibung": ""}


def import_raeume(headers: list, rows: list, mapping: dict, append: bool = False):
    """Importiert Räume. mapping = {app_feld: quelltabellen_index, ...}"""
    if not append:
        # Alle bestehenden Räume ersetzen (Zuordnungen an Optionen bleiben über
        # raum_id bestehen, verweisen dann ggf. ins Leere -> unkritisch, Anzeige
        # zeigt einfach keinen Raumnamen mehr).
        for raum in db.get_all_raeume():
            db.delete_raum(raum["id"])

    for row in rows:
        record = _default_raum()
        for feld, idx in mapping.items():
            if idx is None:
                continue
            try:
                val = row[idx] if idx < len(row) else ""
            except (IndexError, TypeError):
                val = ""
            if feld == "kapazitaet":
                record[feld] = _coerce_int(val)
            else:
                record[feld] = val if val else record[feld]
        # Nur Räume mit echtem Namen anlegen
        if record["name"] and record["name"] != "-":
            db.upsert_raum(record)


def import_projekte(headers: list, rows: list, mapping: dict, append: bool = False):
    """mapping = {app_feld: quelltabellen_index, ...}"""
    if not append:
        db.clear_projekte()

    int_felder = {"nummer", "stufenmin", "stufenmax", "tnmin", "tnmax"}

    # Belegte Nummern (beim Anhängen: bereits vorhandene Optionen). Fehlt die
    # Nummer in der Quelldatei oder kollidiert sie mit einer bereits
    # vergebenen, wird stattdessen fortlaufend weiternummeriert -- so
    # überschreibt "Anhängen" nie bestehende Optionen (upsert_projekt schreibt
    # sonst per ON CONFLICT(nummer) darüber), und eine leere Nummernspalte
    # verhindert den Import einer Zeile nicht mehr.
    vorhandene_nummern = {p["nummer"] for p in db.get_all_projekte()}
    naechste_nummer = max(vorhandene_nummern, default=0) + 1

    for row in rows:
        record = _default_projekt()
        for feld, idx in mapping.items():
            if idx is None:
                continue
            try:
                val = row[idx] if idx < len(row) else ""
            except (IndexError, TypeError):
                val = ""
            if feld in int_felder:
                record[feld] = _coerce_int(val)
            else:
                record[feld] = val if val else record[feld]

        # Leere/Titelzeilen überspringen -- Kriterium ist der Optionsname,
        # nicht mehr die Nummer (die soll fehlen dürfen, siehe oben).
        if not record["projektname"] or record["projektname"] == "-":
            continue

        if record["nummer"] == 0 or record["nummer"] in vorhandene_nummern:
            record["nummer"] = naechste_nummer

        vorhandene_nummern.add(record["nummer"])
        naechste_nummer = max(naechste_nummer, record["nummer"] + 1)

        db.upsert_projekt(record)


# ── TXT-Export ───────────────────────────────────────────────────────────────

def export_txt(filepath: str, rows: list, headers: list, delimiter: str = "\t"):
    with open(filepath, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerow(headers)
        writer.writerows(rows)


# ── Sortier-/Export-Hilfsfunktionen ──────────────────────────────────────────

def _projekt_name(nummer: int, projekte_dict: dict) -> str:
    p = projekte_dict.get(nummer)
    return p["projektname"] if p else str(nummer)


def get_export_data(sort_mode: str, mit_wuenschen: bool = False):
    """
    sort_mode:
      'klasse_name_projekt'
      'klasse_projekt'
      'projekt_klasse_name'
    Gibt (headers, rows) zurück.
    """
    schueler = db.get_all_teilnehmer()
    projekte = {p["nummer"]: p for p in db.get_all_projekte()}

    basis_headers = ["Nachname", "Vorname", "Jgst.", "Abteilung", "Geschlecht", "Projekt-Nr.", "Projektname"]
    wunsch_headers = ["Wunsch 1", "Wunsch 2", "Wunsch 3", "Wunsch 4", "Wunsch 5"]
    headers = basis_headers + (wunsch_headers if mit_wuenschen else [])

    def to_row(s):
        basis = [
            s["nachname"], s["vorname"], s["stufe"], s["stufenzusatz"],
            s["geschlecht"], s["projekt"],
            _projekt_name(s["projekt"], projekte),
        ]
        if mit_wuenschen:
            basis += [s["wunsch_1"], s["wunsch_2"], s["wunsch_3"],
                      s["wunsch_4"], s["wunsch_5"]]
        return basis

    if sort_mode == "klasse_name_projekt":
        schueler.sort(key=lambda s: (s["stufe"], s["stufenzusatz"], s["nachname"], s["vorname"], s["projekt"]))
    elif sort_mode == "klasse_projekt":
        schueler.sort(key=lambda s: (s["stufe"], s["stufenzusatz"], s["projekt"], s["nachname"]))
    elif sort_mode == "projekt_klasse_name":
        schueler.sort(key=lambda s: (s["projekt"], s["stufe"], s["stufenzusatz"], s["nachname"], s["vorname"]))

    rows = [to_row(s) for s in schueler]
    return headers, rows


# ── Excel-Export ─────────────────────────────────────────────────────────────

def export_excel(filepath: str, sort_mode: str, titel: str, jahr: str,
                 mit_wuenschen: bool = False):
    from openpyxl.worksheet.pagebreak import Break

    headers, rows = get_export_data(sort_mode, mit_wuenschen)
    projekte = {p["nummer"]: p for p in db.get_all_projekte()}

    wb = openpyxl.Workbook()
    ws = wb.active

    # Seiteneinrichtung: Hochformat, Inhalt immer auf eine Seitenbreite
    # eingepasst (egal ob mit oder ohne Wünsche, also unabhängig von der
    # Spaltenanzahl), Seitenhöhe nicht begrenzt
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Überschrift
    ueberschrift = f"Projekttage {jahr}: {titel}"
    ws.append([ueberschrift])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.append([])  # Leerzeile

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    current_group = None
    row_num = 3

    def get_group_key(row):
        if sort_mode in ("klasse_name_projekt", "klasse_projekt"):
            return f"Gruppe {row[2]}{row[3]}"
        else:
            import database as _db
            _pl_grp = _db.get_feldkonfig().get("projekt_label", "Projekt")
            return f"{_pl_grp} {row[5]}: {row[6]}"

    # Gruppen-Header + Daten
    for row in rows:
        group = get_group_key(row)
        if group != current_group:
            # Seitenumbruch vor jeder neuen Gruppe (ab der zweiten Gruppe),
            # damit jedes Projekt bzw. jede Klasse auf einer eigenen Seite
            # beginnt
            if current_group is not None:
                ws.row_breaks.append(Break(id=row_num - 1))
                ws.append([])
                row_num += 1

            # Gruppenüberschrift
            ws.append([group])
            cell = ws.cell(row=row_num, column=1)
            cell.font = Font(bold=True, size=12, color="1F3864")
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=len(headers))
            row_num += 1

            # Spaltenheader
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_num, column=col_idx)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
            row_num += 1
            current_group = group

        ws.append(row)
        # Zebrastreifen
        if (row_num % 2) == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                    "solid", fgColor="DCE6F1")
        row_num += 1

    # Spaltenbreiten automatisch
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(filepath)


# ── HTML-Export ──────────────────────────────────────────────────────────────

def export_html(filepath: str, sort_mode: str, titel: str, jahr: str,
                mit_wuenschen: bool = False):
    headers, rows = get_export_data(sort_mode, mit_wuenschen)

    def get_group_key(row):
        if sort_mode in ("klasse_name_projekt", "klasse_projekt"):
            return f"Gruppe {row[2]}{row[3]}"
        else:
            import database as _db
            _pl_grp = _db.get_feldkonfig().get("projekt_label", "Projekt")
            return f"{_pl_grp} {row[5]}: {row[6]}"

    # Gruppen aufbauen
    groups = {}
    for row in rows:
        key = get_group_key(row)
        groups.setdefault(key, []).append(row)

    html_parts = [f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<title>Projekttage {jahr}: {titel}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11pt; margin: 2cm; }}
  h1 {{ color: #1F3864; text-align: center; }}
  h2 {{ color: #2E75B6; margin-top: 0; page-break-before: always; }}
  h2:first-of-type {{ page-break-before: avoid; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 1em; }}
  th {{ background: #4472C4; color: white; padding: 4px 8px; text-align: left; }}
  td {{ padding: 3px 8px; border-bottom: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #DCE6F1; }}
  @media print {{ h2 {{ page-break-before: always; }} }}
</style>
</head>
<body>
<h1>Projekttage {jahr}: {titel}</h1>
"""]

    first = True
    for group, group_rows in groups.items():
        pb = "" if first else ' style="page-break-before:always"'
        html_parts.append(f'<h2{pb}>{group}</h2>\n')
        html_parts.append("<table>\n<tr>")
        for h in headers:
            html_parts.append(f"<th>{h}</th>")
        html_parts.append("</tr>\n")
        for row in group_rows:
            html_parts.append("<tr>")
            for cell in row:
                html_parts.append(f"<td>{cell}</td>")
            html_parts.append("</tr>\n")
        html_parts.append("</table>\n")
        first = False

    html_parts.append("</body>\n</html>")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("".join(html_parts))


# ═══════════════════════════════════════════════════════════════════════════
# Listenexport  (Fensterlisten + Gesamtlisten)
# ═══════════════════════════════════════════════════════════════════════════

def filter_spalten(headers: list, rows: list, kept_indices):
    """
    Reduziert Header- und Zeilenlisten auf die in kept_indices genannten
    Spalten (in der gegebenen Reihenfolge der Originalspalten). Format-agnostisch,
    daher gleichermaßen für xlsx/ods/csv/pdf-Export und HTML-Druck nutzbar.

    kept_indices=None → unverändert zurückgeben (keine Auswahl getroffen).
    Leere Auswahl → ebenfalls unverändert (schützt vor versehentlich leerem Export).
    """
    if kept_indices is None:
        return headers, rows
    idx = [i for i in kept_indices if 0 <= i < len(headers)]
    if not idx:
        return headers, rows
    return (
        [headers[i] for i in idx],
        [[r[i] for i in idx if i < len(r)] for r in rows]
    )


def _hat_wuensch_spalten(gruppen: list) -> bool:
    """True wenn mindestens eine Gruppe Wunsch-Spalten enthält → Querformat."""
    return any(
        any(h.strip().lower().startswith("wunsch") for h in headers)
        for _, headers, _ in gruppen
    )


def _html_gruppen(gruppen: list, kopfzeile: str, seitenumbrueche: bool,
                  datum_fusszeile: bool = False, font_scale: float = 1.0) -> str:
    """
    Baut vollständiges HTML aus. font_scale skaliert alle Schriftgrößen
    (z. B. 1.25 für 25 % größere Darstellung im PDF).
    """
    from datetime import date
    datum_str = date.today().strftime("%d.%m.%Y") if datum_fusszeile else ""

    def pt(size: float) -> str:
        return f"{size * font_scale:.1f}pt"

    css = f"""
    <style>
      body{{font-family:Arial,sans-serif;font-size:{pt(9)};margin:1cm}}
      table{{border-collapse:collapse;width:100%;margin-bottom:6pt}}
      thead{{display:table-header-group}}
      tfoot{{display:table-footer-group}}
      thead .kz td{{
        font-size:{pt(12)};font-weight:bold;color:#1F3864;
        padding:4px 6px;border-bottom:2px solid #1F3864}}
      thead .grp td{{
        font-size:{pt(10)};font-weight:bold;color:#2c3e50;
        padding:2px 6px;border-bottom:1px solid #aaa}}
      thead th{{
        background:#4472C4;color:#ffffff;padding:3px 6px;
        font-size:{pt(8.5)};text-align:left;font-weight:bold}}
      td{{padding:2px 6px;border-bottom:1px solid #ddd;font-size:{pt(8.5)}}}
      tr:nth-child(even) td{{background:#f0f4f8}}
      .pgbrk{{page-break-after:always}}
      .fuss td{{font-size:{pt(7.5)};color:#666;padding-top:4px;
               border-top:1px solid #ccc}}
    </style>"""
    parts = [f"<!DOCTYPE html><html lang='de'><head><meta charset='UTF-8'>{css}</head><body>"]
    for gi, (gruppenname, headers, rows) in enumerate(gruppen):
        last = gi == len(gruppen) - 1
        pgcls = "" if last or not seitenumbrueche else " class='pgbrk'"
        parts.append(f"<table{pgcls}><thead>")
        if kopfzeile:
            parts.append(
                f"<tr class='kz'>"
                f"<td colspan='{len(headers)}'>{kopfzeile}</td></tr>"
            )
        if gruppenname:
            parts.append(
                f"<tr class='grp'>"
                f"<td colspan='{len(headers)}'>{gruppenname}</td></tr>"
            )
        parts.append("<tr>")
        for h in headers:
            parts.append(f"<th>{h}</th>")
        parts.append("</tr></thead>")
        if datum_fusszeile:
            parts.append(
                f"<tfoot><tr class='fuss'>"
                f"<td colspan='{len(headers)}'>Stand: {datum_str}</td>"
                f"</tr></tfoot>"
            )
        parts.append("<tbody>")
        for row in rows:
            parts.append("<tr>")
            for val in row:
                if isinstance(val, db.Geaendert):
                    parts.append(f"<td style='background:{db.HERVORHEBUNG_GEAENDERT_HEX}'>{val}</td>")
                elif isinstance(val, db.Geist):
                    parts.append(f"<td style='background:{db.HERVORHEBUNG_GEIST_HEX}'>{val}</td>")
                else:
                    parts.append(f"<td>{val}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table>")
    parts.append("</body></html>")
    return "".join(parts)


def export_gruppen_xlsx(path: str, gruppen: list, kopfzeile: str,
                         seitenumbrueche: bool, datum_fusszeile: bool = False):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.pagebreak import Break

    querformat = _hat_wuensch_spalten(gruppen)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.page_setup.orientation = (ws.ORIENTATION_LANDSCAPE
                                  if querformat else ws.ORIENTATION_PORTRAIT)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    cur_row = 1
    if kopfzeile:
        ws.cell(row=cur_row, column=1, value=kopfzeile).font = Font(bold=True, size=13)
        ws.oddHeader.center.text = kopfzeile
        ws.evenHeader.center.text = kopfzeile
        cur_row += 1

    header_fill = PatternFill("solid", fgColor="4472C4")

    for gi, (gruppenname, headers, rows) in enumerate(gruppen):
        if gruppenname:
            ws.cell(row=cur_row, column=1,
                    value=gruppenname).font = Font(bold=True, size=11)
            cur_row += 1
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=cur_row, column=ci, value=h)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
        cur_row += 1
        fill_geaendert = PatternFill(
            "solid", fgColor=db.HERVORHEBUNG_GEAENDERT_HEX.lstrip("#"))
        fill_geist = PatternFill(
            "solid", fgColor=db.HERVORHEBUNG_GEIST_HEX.lstrip("#"))
        for row in rows:
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=cur_row, column=ci, value=val)
                # Nachbearbeitungsmodus-Marker farblich hervorheben
                if isinstance(val, db.Geaendert):
                    cell.fill = fill_geaendert
                elif isinstance(val, db.Geist):
                    cell.fill = fill_geist
            cur_row += 1
        if seitenumbrueche and gi < len(gruppen) - 1:
            ws.row_breaks.append(Break(id=cur_row - 1))
        cur_row += 1

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max_len + 4, 45)

    if datum_fusszeile:
        from datetime import date
        datum_str = date.today().strftime("%d.%m.%Y")
        ws.oddFooter.left.text = f"Stand: {datum_str}"
        ws.evenFooter.left.text = f"Stand: {datum_str}"

    wb.save(path)


def export_gruppen_ods(path: str, gruppen: list, kopfzeile: str,
                       seitenumbrueche: bool = False, datum_fusszeile: bool = False):
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import (Style, TextProperties, TableCellProperties,
                            TableRowProperties, PageLayout,
                            PageLayoutProperties, MasterPage, Footer)
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    from odf.namespaces import STYLENS

    querformat = _hat_wuensch_spalten(gruppen)

    doc = OpenDocumentSpreadsheet()

    # Seitenausrichtung + ggf. Platz für Fußzeile
    bottom_margin = "1.5cm" if datum_fusszeile else "1cm"
    pl = PageLayout(name="Seite")
    pl.addElement(PageLayoutProperties(
        printorientation="landscape" if querformat else "portrait",
        margintop="1cm", marginbottom=bottom_margin,
        marginleft="1cm", marginright="1cm"
    ))
    doc.automaticstyles.addElement(pl)
    mp = MasterPage(name="Standard", pagelayoutname="Seite")

    if datum_fusszeile:
        from datetime import date as _date
        datum_str = _date.today().strftime("%d.%m.%Y")
        footer_elem = Footer()
        footer_p = P(text=f"Stand: {datum_str}")
        footer_elem.addElement(footer_p)
        mp.addElement(footer_elem)

    doc.masterstyles.addElement(mp)

    def _add_cell_style(name, **props):
        s = Style(name=name, family="table-cell")
        tp = {k: v for k, v in props.items() if k in ("fontweight", "color")}
        if tp:
            s.addElement(TextProperties(**tp))
        if "backgroundcolor" in props:
            s.addElement(TableCellProperties(
                backgroundcolor=props["backgroundcolor"]))
        doc.styles.addElement(s)

    _add_cell_style("Bold",   fontweight="bold")
    _add_cell_style("Header", fontweight="bold",
                    color="#ffffff", backgroundcolor="#4472C4")
    _add_cell_style("Geaendert", backgroundcolor=db.HERVORHEBUNG_GEAENDERT_HEX)
    _add_cell_style("Geist",     backgroundcolor=db.HERVORHEBUNG_GEIST_HEX)

    brk = Style(name="PageBreak", family="table-row")
    brk.addElement(TableRowProperties(breakbefore="page"))
    doc.automaticstyles.addElement(brk)

    # Tabellenstil, der explizit auf unsere MasterPage zeigt —
    # erst dadurch greift die Fußzeile auch in LibreOffice Calc
    table = Table(name="Export")
    if datum_fusszeile:
        table.setAttrNS(STYLENS, "master-page-name", "Standard")

    def _row(*vals, cell_style=None, row_style=None):
        tr = TableRow(stylename=row_style) if row_style else TableRow()
        for v in vals:
            # Ohne expliziten Stil bekommen Nachbearbeitungsmodus-Marker
            # ihren Hervorhebungsstil (pro Zelle).
            stil = cell_style
            if stil is None:
                if isinstance(v, db.Geaendert):
                    stil = "Geaendert"
                elif isinstance(v, db.Geist):
                    stil = "Geist"
            tc = (TableCell(stylename=stil, valuetype="string")
                  if stil else TableCell(valuetype="string"))
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    if kopfzeile:
        table.addElement(_row(kopfzeile, cell_style="Bold"))
        table.addElement(TableRow())

    for gi, (gruppenname, headers, rows) in enumerate(gruppen):
        row_style = "PageBreak" if seitenumbrueche and gi > 0 else None
        if gruppenname:
            table.addElement(_row(gruppenname, cell_style="Bold",
                                   row_style=row_style))
            row_style = None
        table.addElement(_row(*headers, cell_style="Header",
                               row_style=row_style))
        for row in rows:
            table.addElement(_row(*row))
        table.addElement(TableRow())

    doc.spreadsheet.addElement(table)
    doc.save(path)


def export_gruppen_csv(path: str, gruppen: list, kopfzeile: str):
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        if kopfzeile:
            w.writerow([kopfzeile])
            w.writerow([])
        for gruppenname, headers, rows in gruppen:
            if gruppenname:
                w.writerow([gruppenname])
            w.writerow(headers)
            for row in rows:
                w.writerow(row)
            w.writerow([])


def export_gruppen_pdf(path: str, gruppen: list, kopfzeile: str,
                        seitenumbrueche: bool, datum_fusszeile: bool = False):
    from PyQt6.QtPrintSupport import QPrinter
    from PyQt6.QtGui import (QTextDocument, QPageLayout, QPageSize,
                              QPainter, QFont, QColor)
    from PyQt6.QtCore import QMarginsF, QSizeF, QRectF, Qt

    from datetime import date as _date
    datum_str = _date.today().strftime("%d.%m.%Y") if datum_fusszeile else ""

    querformat = _hat_wuensch_spalten(gruppen)
    orientierung = (QPageLayout.Orientation.Landscape
                    if querformat else QPageLayout.Orientation.Portrait)

    # HTML ohne Datum — wird per QPainter auf jede Seite gezeichnet
    html = _html_gruppen(gruppen, kopfzeile, seitenumbrueche, datum_fusszeile=False)

    # HTML mit 25 % größeren Schriften → zuverlässige Skalierung unabhängig
    # von QPrinter-Modus und Koordinatensystem-Eigenheiten
    html = _html_gruppen(gruppen, kopfzeile, seitenumbrueche,
                         datum_fusszeile=False, font_scale=1.25)

    printer = QPrinter(QPrinter.PrinterMode.ScreenResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(path)
    layout = QPageLayout(
        QPageSize(QPageSize.PageSizeId.A4),
        orientierung,
        QMarginsF(15, 10, 15, 10)
    )
    printer.setPageLayout(layout)

    if not datum_fusszeile:
        # Kein Footer nötig → einfaches doc.print()
        doc = QTextDocument()
        doc.setHtml(html)
        getattr(doc, 'print')(printer)
        return

    # Mit Footer: seitenweises Rendering per QPainter
    page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
    page_w = page_rect.width()
    page_h = page_rect.height()

    doc = QTextDocument()
    doc.setHtml(html)
    doc.setPageSize(QSizeF(page_w, page_h))

    painter = QPainter(printer)
    n_pages = doc.pageCount()

    for page in range(n_pages):
        if page > 0:
            printer.newPage()

        # Clip verhindert Bleed-Through des nächsten Gruppenheaders
        painter.setClipRect(QRectF(0, 0, page_w, page_h))
        painter.save()
        painter.translate(0.0, -page * page_h)
        doc.drawContents(painter)
        painter.restore()

        # Datum links unten
        painter.setClipping(False)
        painter.save()
        font = QFont("Arial")
        font.setPointSize(8)
        painter.setFont(font)
        painter.setPen(QColor(100, 100, 100))
        fm = painter.fontMetrics()
        line_h = fm.height()
        y_line = int(page_h - line_h * 2.2)
        painter.drawLine(0, y_line, int(page_w), y_line)
        footer_rect = QRectF(0, y_line + 2, page_w, line_h * 1.5)
        painter.drawText(footer_rect,
                         Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                         f"Stand: {datum_str}")
        painter.restore()

    painter.end()


def export_gruppen(path: str, fmt: str, gruppen: list,
                   kopfzeile: str = "", seitenumbrueche: bool = False,
                   datum_fusszeile: bool = False):
    """Einheitlicher Einstiegspunkt für alle Exportformate."""
    if fmt == "xlsx":
        export_gruppen_xlsx(path, gruppen, kopfzeile, seitenumbrueche, datum_fusszeile)
    elif fmt == "ods":
        export_gruppen_ods(path, gruppen, kopfzeile, seitenumbrueche, datum_fusszeile)
    elif fmt == "csv":
        export_gruppen_csv(path, gruppen, kopfzeile)
    elif fmt == "pdf":
        export_gruppen_pdf(path, gruppen, kopfzeile, seitenumbrueche, datum_fusszeile)
    else:
        raise ValueError(f"Unbekanntes Format: {fmt}")




def get_gesamtliste_nach_klassen(mit_wuenschen: bool = True) -> list:
    """[(gruppenname, headers, rows), ...] sortiert nach Stufe/Zusatz/Name."""
    from itertools import groupby
    k = db.get_feldkonfig()
    sl  = k.get("stufe_label",        "Gruppenbereich")
    zl  = k.get("stufenzusatz_label", "Gruppenzusatz")
    pl  = k.get("projekt_label",       "Projekt")

    alle = db.get_all_teilnehmer()
    alle.sort(key=lambda s: (
        int(s["stufe"]) if str(s["stufe"]).isdigit() else 9999,
        str(s["stufenzusatz"] or ""),
        str(s["nachname"] or ""),
        str(s["vorname"]  or ""),
    ))

    if mit_wuenschen:
        hdrs = ["Name", sl, zl,
                "Wunsch 1", "Wunsch 2", "Wunsch 3", "Wunsch 4", "Wunsch 5",
                f"{pl}-Nr."]
    else:
        hdrs = ["Name", sl, zl, f"{pl}-Nr."]

    gruppen = []
    for (stufe, zusatz), grp in groupby(
            alle, key=lambda s: (s["stufe"], s["stufenzusatz"])):
        grp = list(grp)
        label = (f"{stufe}{zusatz}"
                 if zusatz and str(zusatz) not in ("-", "")
                 else f"{sl} {stufe}")
        rows = []
        for s in grp:
            # Bei aktivem Bearbeitungsmodus und Abweichung den Marker anzeigen,
            # sonst wie bisher (leer bei unzugeteilt).
            az = db.zuteilung_anzeige(s)
            p = az if isinstance(az, db.Geaendert) else (s["projekt"] if s["projekt"] else "")
            if mit_wuenschen:
                rows.append([
                    f"{s['nachname']}, {s['vorname']}",
                    s["stufe"], s["stufenzusatz"],
                    s["wunsch_1"] or "", s["wunsch_2"] or "",
                    s["wunsch_3"] or "", s["wunsch_4"] or "",
                    s["wunsch_5"] or "", p,
                ])
            else:
                rows.append([
                    f"{s['nachname']}, {s['vorname']}",
                    s["stufe"], s["stufenzusatz"], p,
                ])
        gruppen.append((label, hdrs, rows))
    return gruppen


def get_gesamtliste_nach_projekten(mit_wuenschen: bool = True) -> list:
    """[(gruppenname, headers, rows), ...] sortiert nach Projekt/Name."""
    from itertools import groupby
    k = db.get_feldkonfig()
    sl  = k.get("stufe_label",        "Gruppenbereich")
    zl  = k.get("stufenzusatz_label", "Gruppenzusatz")

    alle_tn   = db.get_all_teilnehmer()
    alle_proj = {p["nummer"]: p for p in db.get_all_projekte()}

    alle_tn.sort(key=lambda s: (
        s["projekt"] or 0,
        str(s["nachname"] or ""),
        str(s["vorname"]  or ""),
    ))

    if mit_wuenschen:
        hdrs = ["Name", sl, zl,
                "Wunsch 1", "Wunsch 2", "Wunsch 3", "Wunsch 4", "Wunsch 5"]
    else:
        hdrs = ["Name", sl, zl]

    gruppen = []
    for projekt_nr, grp in groupby(alle_tn, key=lambda s: s["projekt"]):
        grp = list(grp)
        if not projekt_nr:
            label = "0 – (noch nicht zugeteilt)"
        elif projekt_nr in alle_proj:
            label = f"{projekt_nr}: {alle_proj[projekt_nr]['projektname']}"
        else:
            label = str(projekt_nr)
        rows = []
        for s in grp:
            if mit_wuenschen:
                rows.append([
                    f"{s['nachname']}, {s['vorname']}",
                    s["stufe"], s["stufenzusatz"],
                    s["wunsch_1"] or "", s["wunsch_2"] or "",
                    s["wunsch_3"] or "", s["wunsch_4"] or "",
                    s["wunsch_5"] or "",
                ])
            else:
                rows.append([
                    f"{s['nachname']}, {s['vorname']}",
                    s["stufe"], s["stufenzusatz"],
                ])
        gruppen.append((label, hdrs, rows))
    return gruppen


def export_gruppen_separat(pfad_oder_ordner: str, fmt: str, gruppen: list,
                            kopfzeile: str, datum_fusszeile: bool,
                            als_zip: bool = True):
    """
    Exportiert jede Gruppe als separate Datei.

    als_zip=True:  Alle Dateien in ein ZIP-Archiv (pfad_oder_ordner = .zip-Pfad)
    als_zip=False: Alle Dateien direkt in den angegebenen Ordner schreiben
    """
    import zipfile, io, tempfile, os, pathlib

    suffix = f".{fmt}"
    safe_name = lambda n: "".join(c if c.isalnum() or c in " _-" else "_" for c in n)

    if als_zip:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for gruppenname, headers, rows in gruppen:
                fname = safe_name(gruppenname or "Gruppe") + suffix
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tf:
                    tmp = tf.name
                try:
                    export_gruppen(tmp, fmt, [(gruppenname, headers, rows)],
                                   kopfzeile, seitenumbrueche=False,
                                   datum_fusszeile=datum_fusszeile)
                    zf.write(tmp, fname)
                finally:
                    os.unlink(tmp)
        with open(pfad_oder_ordner, "wb") as f:
            f.write(buf.getvalue())
    else:
        ordner = pathlib.Path(pfad_oder_ordner)
        ordner.mkdir(parents=True, exist_ok=True)
        for gruppenname, headers, rows in gruppen:
            fname = safe_name(gruppenname or "Gruppe") + suffix
            ziel  = str(ordner / fname)
            export_gruppen(ziel, fmt, [(gruppenname, headers, rows)],
                           kopfzeile, seitenumbrueche=False,
                           datum_fusszeile=datum_fusszeile)
