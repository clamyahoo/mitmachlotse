"""
Wiederverwendbares Fenster zur Anzeige von Listen (Wunschauswertung,
Projektteilnehmerlisten, Klassenlisten) mit Export- und Druckfunktion.
"""

import os
import database as db
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QMessageBox, QAbstractItemView, QMenu
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QAction, QKeySequence, QPalette, QColor
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog
from PyQt6.QtGui import QTextDocument

import importexport as ie


def _erzwinge_lesbare_selektion(table: QTableWidget):
    """
    Erzwingt einen lesbaren Kontrast bei markierten Zeilen, unabhängig vom
    Betriebssystem-Theme (siehe gleichnamige Funktion in hauptfenster.py).
    """
    palette = table.palette()
    palette.setColor(QPalette.ColorGroup.Active,
                     QPalette.ColorRole.Highlight, QColor("#2980b9"))
    palette.setColor(QPalette.ColorGroup.Active,
                     QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorGroup.Inactive,
                     QPalette.ColorRole.Highlight, QColor("#6b95b5"))
    palette.setColor(QPalette.ColorGroup.Inactive,
                     QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
    table.setPalette(palette)


class ListenFenster(QDialog):
    """
    Generisches, nicht-modales Fenster zur Anzeige einer Liste mit
    Export (TXT/Excel/HTML) und Druckfunktion.

    headers: Liste von Spaltenüberschriften
    rows:    Liste von Listen/Tupeln mit den Zeilenwerten
    titel:   Fenstertitel und Überschrift in Export/Druck
    row_ids: optionale Liste von Schüler-IDs (parallel zu rows), ermöglicht
             die "Zuteilen"-Funktion direkt aus dem Fenster heraus
    zuteilen_callback: optionale Funktion(schueler_id), die beim Klick auf
             "Zuteilen" für die ausgewählte Zeile aufgerufen wird; soll
             nach erfolgter Zuteilung True zurückgeben, damit die Liste
             aktualisiert werden kann
    wunsch_bearbeiten_callback: optionale Funktion(schueler_id, wunsch_rang)
             -- wird aufgerufen, wenn auf eine Zelle einer Spalte
             "Wunsch 1".."Wunsch 5" doppelgeklickt wird (z. B. um den
             Wunsch über einen Auswahldialog direkt zu korrigieren).
             wunsch_rang ist 1-basiert (1 = "Wunsch 1").
    """

    def __init__(self, titel: str, headers: list, rows: list, parent=None,
                 row_ids: list = None, zuteilen_callback=None,
                 wunsch_bearbeiten_callback=None, details_callback=None):
        super().__init__(parent)
        self.setWindowTitle(titel)
        self.setMinimumSize(800, 500)
        self.setModal(False)  # Mehrere Listenfenster gleichzeitig offen möglich

        # Als vollwertiges, eigenständiges Fenster behandeln (mit
        # Minimieren-/Maximieren-Schaltflächen und eigenem Eintrag in der
        # Taskleiste). Das ist auf den meisten Betriebssystemen (Windows,
        # GNOME, KDE) Voraussetzung dafür, dass der Fenstermanager das
        # Fenster per Klick in den Vordergrund holen und per Ziehen an den
        # Bildschirmrand andocken ("Snapping") lässt -- ein reiner QDialog
        # wird von manchen Fenstermanagern eingeschränkter behandelt.
        self.setWindowFlags(
            Qt.WindowType.Window
            | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint
            | Qt.WindowType.WindowCloseButtonHint
        )

        self._titel = titel
        self._headers = headers
        self._rows = rows
        self._row_ids = row_ids
        self._zuteilen_callback = zuteilen_callback
        self._wunsch_bearbeiten_callback = wunsch_bearbeiten_callback
        self._details_callback = details_callback
        self._last_dir = os.path.expanduser("~")

        layout = QVBoxLayout(self)

        # Kopfzeile mit Titel und Trefferanzahl
        kopf_layout = QHBoxLayout()
        titel_label = QLabel(f"<b>{titel}</b>")
        titel_label.setStyleSheet("font-size: 15px;")
        kopf_layout.addWidget(titel_label)
        kopf_layout.addStretch()
        self.lbl_anzahl = QLabel(f"{len(rows)} Einträge")
        kopf_layout.addWidget(self.lbl_anzahl)
        layout.addLayout(kopf_layout)

        # Tabelle
        self.table = QTableWidget(len(rows), len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        _erzwinge_lesbare_selektion(self.table)
        self._setze_stretch_spalte(headers)
        if wunsch_bearbeiten_callback is not None and row_ids is not None:
            self.table.cellDoubleClicked.connect(self._on_wunsch_doppelklick)
            hinweis = QLabel(
                "Doppelklick auf ein Wunschfeld öffnet die Auswahl der "
                f"zulässigen {db.get_feldkonfig().get('projekt_label','Optionen')} für diese Person."
            )
            hinweis.setStyleSheet("color: #555; font-size: 11px;")
            layout.addWidget(hinweis)
        elif zuteilen_callback is not None and row_ids is not None:
            # Kein Wunsch-Callback, aber Zuteilungs-Callback:
            # Doppelklick auf Projekt-Spalte öffnet Zuweisungsdialog
            self.table.cellDoubleClicked.connect(self._on_projekt_doppelklick)
        self._fill_table(rows)
        layout.addWidget(self.table)

        # Buttons
        btn_layout = QHBoxLayout()

        if row_ids is not None and zuteilen_callback is not None:
            _pl = db.get_feldkonfig().get("projekt_label", "Projekt")
            btn_zuteilen = QPushButton(f"{_pl} fix zuweisen")
            btn_zuteilen.clicked.connect(self._zuteilen)
            btn_layout.addWidget(btn_zuteilen)
            # Tastenkürzel – identisch mit dem Hauptfenster
            sc_zuteilen = QAction(self)
            sc_zuteilen.setShortcut(QKeySequence("Ctrl+Shift+F"))
            sc_zuteilen.triggered.connect(self._zuteilen)
            self.addAction(sc_zuteilen)

        if details_callback is not None:
            _kf_lf = db.get_feldkonfig()
            _pl_lf = _kf_lf.get("projekt_label", "Projekt")
            _f_lf  = db.get_label_formen(_pl_lf)
            btn_details = QPushButton(f"Details {_f_lf['dat_dem']}")
            btn_details.clicked.connect(lambda: details_callback())
            btn_layout.addWidget(btn_details)

        btn_export = QPushButton("Exportieren")
        btn_export.clicked.connect(self._export)
        btn_print = QPushButton("Drucken")
        btn_print.clicked.connect(self._print)
        btn_preview = QPushButton("Druckvorschau")
        btn_preview.clicked.connect(self._print_preview)
        btn_close = QPushButton("Schließen")
        btn_close.clicked.connect(self.close)

        btn_layout.addWidget(btn_export)
        btn_layout.addWidget(btn_print)
        btn_layout.addWidget(btn_preview)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

    def _on_wunsch_doppelklick(self, row: int, col: int):
        """Reagiert auf Doppelklick in einer Wunsch-Spalte (Wunsch 1-5)
        und ruft den Bearbeiten-Callback mit (schueler_id, wunsch_rang) auf."""
        if self._row_ids is None or row >= len(self._row_ids):
            return
        if col >= len(self._headers):
            return
        header = self._headers[col].strip().lower()
        if not header.startswith("wunsch "):
            return
        try:
            wunsch_rang = int(header.replace("wunsch", "").strip())
        except ValueError:
            return
        if not (1 <= wunsch_rang <= 5):
            return
        schueler_id = self._row_ids[row]
        if self._wunsch_bearbeiten_callback:
            self._wunsch_bearbeiten_callback(schueler_id, wunsch_rang)

    def _on_projekt_doppelklick(self, row: int, col: int):
        """Doppelklick auf beliebige Zelle → Zuweisungsdialog öffnen
        (wird in Listen ohne Wunsch-Callback verwendet, z. B. Projektteilnehmerliste)."""
        if self._row_ids is None or row >= len(self._row_ids):
            return
        self.table.selectRow(row)
        self._zuteilen()

    # Spaltennamen, die typischerweise lange, variable Textinhalte haben
    # und sich daher strecken sollen, um den verfügbaren Platz zu nutzen
    _STRETCH_KANDIDATEN = ["name", "projektname"]

    def _setze_stretch_spalte(self, headers: list):
        """
        Lässt die Spalte mit langen, variablen Textinhalten (i. d. R.
        "Name") sich strecken, statt fest die Spalte an Index 1 zu nehmen
        -- die Spaltenreihenfolge unterscheidet sich je nach Liste, "Name"
        steht zwar meist an Index 0, aber das soll nicht angenommen
        werden müssen.
        """
        if not headers:
            return
        header_lower = [h.strip().lower() for h in headers]
        ziel_index = None
        for kandidat in self._STRETCH_KANDIDATEN:
            if kandidat in header_lower:
                ziel_index = header_lower.index(kandidat)
                break
        if ziel_index is None:
            ziel_index = 0  # Fallback: erste Spalte
        self.table.horizontalHeader().setSectionResizeMode(
            ziel_index, QHeaderView.ResizeMode.Stretch
        )

    def _zuteilen(self):
        row = self.table.currentRow()
        if row < 0 or self._row_ids is None or row >= len(self._row_ids):
            QMessageBox.information(
                self, "Hinweis", "Bitte zuerst eine Zeile in der Tabelle auswählen."
            )
            return
        schueler_id = self._row_ids[row]
        if self._zuteilen_callback:
            self._zuteilen_callback(schueler_id)
        # Fokus nach der Zuteilung im Listenfenster belassen
        self.raise_()
        self.activateWindow()

    def aktualisiere_daten(self, headers: list, rows: list, row_ids: list = None):
        """Erlaubt es, die Tabelleninhalte nach einer Änderung neu zu laden."""
        self._headers = headers
        self._rows = rows
        self._row_ids = row_ids
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self._setze_stretch_spalte(headers)
        self._fill_table(rows)

    def _fill_table(self, rows: list):
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                if c != 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(r, c, item)
        self._setze_spaltenbreiten()
        self.lbl_anzahl.setText(f"{len(rows)} Einträge")

    # Spalten, die von Natur aus nur sehr kurze Werte enthalten (Zahlen,
    # einzelne Buchstaben, kurze Status-Wörter) -- diese bekommen eine feste,
    # kompakte Breite statt sich an resizeColumnsToContents() zu orientieren,
    # das bei kurzen Headern trotzdem unnötig viel Platz beanspruchen kann.
    _SCHMALE_SPALTEN = {
        "jgst.": 50,
        "klassenzusatz": 95,
        "geschl.": 65,
        "geschlecht": 90,
        "wunsch 1": 70, "wunsch 2": 70, "wunsch 3": 70,
        "wunsch 4": 70, "wunsch 5": 70,
        "projekt": 70,
        "projekt-nr.": 80,
        "wunschrang": 85,
        "zugeteilt?": 80,
        "anzahl wünsche": 110,
    }

    def _setze_spaltenbreiten(self):
        """
        Setzt für bekannte schmale Spalten (Jgst., Geschlecht, Wunsch-
        Spalten usw.) eine kompakte feste Breite, lässt breitere
        Inhaltsspalten (Name, Projektname, Wunschrang erhalten ...) sich
        an ihrem Inhalt orientieren bzw. strecken. So verteilt sich der
        verfügbare Platz sinnvoller, statt dass kurze Spalten unnötig
        breit geraten.
        """
        self.table.resizeColumnsToContents()
        for c, header in enumerate(self._headers):
            breite = self._SCHMALE_SPALTEN.get(header.strip().lower())
            if breite is not None:
                self.table.setColumnWidth(c, breite)
        self._passe_fensterbreite_an()

    def _passe_fensterbreite_an(self):
        """
        Vergrößert das Fenster bei Bedarf so, dass alle Spalten ohne
        horizontales Scrollen sichtbar sind -- bei vielen Spalten (z. B.
        Wunsch 1-5 + Projekt + Wunschrang erhalten) reicht die feste
        Mindestbreite sonst nicht aus.
        """
        gesamtbreite = self.table.verticalHeader().width()
        for c in range(self.table.columnCount()):
            gesamtbreite += self.table.columnWidth(c)
        # Rahmen/Scrollbar-Puffer und Außenabstände des Fensters
        gesamtbreite += 60

        bildschirm = self.screen() if hasattr(self, "screen") and self.screen() else None
        max_breite = bildschirm.availableGeometry().width() - 80 if bildschirm else 1800
        ziel_breite = max(800, min(gesamtbreite, max_breite))

        aktuelle_groesse = self.size()
        if ziel_breite > aktuelle_groesse.width():
            self.resize(ziel_breite, aktuelle_groesse.height())
        # Mindestbreite ebenfalls anheben, damit die Spalten beim manuellen
        # Verkleinern nicht sofort wieder abgeschnitten werden
        self.setMinimumWidth(min(ziel_breite, 800))

    # ── Export ───────────────────────────────────────────────────────────────

    def _export(self):
        from dialoge import FensterExportDialog
        hat_wuensche = any(h.strip().lower().startswith("wunsch")
                           for h in self._headers)
        dlg = FensterExportDialog(hat_wuensche=hat_wuensche, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        fmt = dlg.get_format()
        mit_wuenschen = dlg.get_mit_wuenschen()

        # Spalten filtern
        headers, rows = ie.filter_wunsch_spalten(
            self._headers, self._rows, mit_wuenschen
        )
        # Fixiert-Spalte raus (wie beim Drucken)
        fix_idx = [i for i, h in enumerate(headers)
                   if h.strip().lower() == "fixiert"]
        if fix_idx:
            fi = fix_idx[0]
            headers = [h for i, h in enumerate(headers) if i != fi]
            rows = [[c for i, c in enumerate(r) if i != fi] for r in rows]

        ext_map = {"xlsx": ".xlsx", "ods": ".ods",
                   "csv": ".csv", "pdf": ".pdf"}
        pfad, _ = QFileDialog.getSaveFileName(
            self, "Liste exportieren",
            os.path.join(self._last_dir, self._dateiname().replace(".txt", ext_map[fmt])),
            f"{fmt.upper()}-Datei (*{ext_map[fmt]})"
        )
        if not pfad:
            return
        if not pfad.lower().endswith(ext_map[fmt]):
            pfad += ext_map[fmt]
        self._last_dir = os.path.dirname(pfad)

        try:
            ie.export_gruppen(
                pfad, fmt,
                [(self._titel, headers, rows)],
                kopfzeile=self._titel,
                seitenumbrueche=False
            )
            QMessageBox.information(self, "Export erfolgreich",
                                    f"Datei gespeichert:\n{pfad}")
        except Exception as e:
            QMessageBox.critical(self, "Fehler beim Export", str(e))

    def _dateiname(self) -> str:
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in self._titel)
        return safe.strip().replace(" ", "_") + ".txt"

    def _export_excel(self, pfad: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active

        # Seiteneinrichtung: Hochformat, Inhalt auf eine Seitenbreite
        # eingepasst (konsistent mit dem Haupt-Datenexport)
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.append([self._titel])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self._headers))
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(self._headers)
        header_fill = PatternFill("solid", fgColor="4472C4")
        for col_idx in range(1, len(self._headers) + 1):
            c = ws.cell(row=3, column=col_idx)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
        for row in self._rows:
            ws.append(list(row))
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = None
            for cell in col:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        wb.save(pfad)

    def _export_html(self, pfad: str):
        html = self._build_html_document()
        with open(pfad, "w", encoding="utf-8") as f:
            f.write(html)

    # ── Druck ────────────────────────────────────────────────────────────────

    def _build_html_document(self) -> str:
        # "Fixiert"-Spalte im Druck weglassen
        druck_cols = [i for i, h in enumerate(self._headers)
                      if h.strip().lower() != "fixiert"]
        headers = [self._headers[i] for i in druck_cols]
        rows    = [[r[i] for i in druck_cols if i < len(r)]
                   for r in self._rows]

        parts = [f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<title>{self._titel}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 10pt; margin: 1.5cm; }}
  h1 {{ color: #1F3864; font-size: 14pt; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #4472C4; color: white; padding: 4px 8px; text-align: left; }}
  td {{ padding: 3px 8px; border-bottom: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #f0f4f8; }}
</style>
</head>
<body>
<h1>{self._titel}</h1>
<table>
<tr>"""]
        for h in headers:
            parts.append(f"<th>{h}</th>")
        parts.append("</tr>\n")
        for row in rows:
            parts.append("<tr>")
            for val in row:
                parts.append(f"<td>{val}</td>")
            parts.append("</tr>\n")
        parts.append("</table>\n</body>\n</html>")
        return "".join(parts)

    def _print(self):
        doc = QTextDocument()
        doc.setHtml(self._build_html_document())
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            doc.print(printer)

    def _print_preview(self):
        doc = QTextDocument()
        doc.setHtml(self._build_html_document())
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        preview = QPrintPreviewDialog(printer, self)
        preview.paintRequested.connect(lambda p: doc.print(p))
        preview.exec()
